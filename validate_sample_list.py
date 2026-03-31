#!/usr/bin/env python3
"""
Validate the first sheet of an RPPA sample list xlsx file.

Checks performed on every data row:
  1. Whitespace   – leading/trailing whitespace in SampleName, Name in print map, SampleID
  2. SpecialChars – characters other than [A-Za-z0-9_-] in the same three columns
  3. CalPrefix    – entries starting with "Cal" but not "Cal_"
  4. Concentration – values in Normalization / Actual Conc. that are not 0.5 or "no normalization"

Outputs:
  - A validation report (.xlsx) with pass/fail per check and an overall Result column
  - A cleaned sample list (.xlsx) with whitespace stripped from the three name columns
"""

import glob
import re
import shutil
import sys
import os
import openpyxl

WORKDIR = os.path.dirname(os.path.abspath(__file__))

INPUT_PATTERN = "RPPA*sample list.xlsx"


def find_input_file():
    """Find exactly one file matching RPPA*sample list.xlsx in the working directory."""
    matches = glob.glob(os.path.join(WORKDIR, INPUT_PATTERN))
    # Exclude Excel temp files (~$...)
    matches = [m for m in matches if not os.path.basename(m).startswith("~$")]
    if len(matches) == 0:
        print(f"ERROR: No file matching '{INPUT_PATTERN}' found in {WORKDIR}")
        sys.exit(1)
    if len(matches) > 1:
        print(f"ERROR: Multiple files matching '{INPUT_PATTERN}' found in {WORKDIR}:")
        for m in matches:
            print(f"  {m}")
        sys.exit(1)
    return matches[0]


def derive_output_paths(input_path):
    """Derive report and cleaned file paths from the input filename."""
    basename = os.path.basename(input_path)
    # e.g. "RPPA0065_sample list.xlsx" -> stem = "RPPA0065_sample list"
    stem = os.path.splitext(basename)[0]
    # Replace spaces with underscores for output filenames
    stem_safe = stem.replace(" ", "_")
    directory = os.path.dirname(input_path)
    report = os.path.join(directory, f"{stem_safe}_validation_report.xlsx")
    cleaned = os.path.join(directory, f"{stem_safe}_cleaned.xlsx")
    return report, cleaned

EXPECTED_HEADERS = [
    "#",
    "SampleName",
    "Name in print map",
    "SampleID",
    "Treatment Group",
    "Sample Type",
    "Normalization Conc.(ug/ul)",
    "Actual Conc.(ug/ul)",
]

NAME_COLS = [1, 2, 3]  # 0-indexed positions of SampleName, Name in print map, SampleID
CONC_COLS = [6, 7]      # 0-indexed positions of the two concentration columns

VALID_CHAR_RE = re.compile(r"^[A-Za-z0-9_\-]+$")


def check_whitespace(value):
    """Return True if the string value has leading or trailing whitespace."""
    s = str(value)
    return s != s.strip()


def check_special_chars(value):
    """Return True if the value contains disallowed characters (anything not [A-Za-z0-9_-])."""
    return not VALID_CHAR_RE.match(str(value))


def check_cal_prefix(value):
    """Return True (fail) if value starts with 'Cal' but not 'Cal_'."""
    s = str(value)
    if s.startswith("Cal"):
        return not s.startswith("Cal_")
    return False


def check_concentration(value):
    """Return True (fail) if value is not 0.5, "0.5", or "no normalization" (case-sensitive)."""
    if isinstance(value, (int, float)):
        return value != 0.5
    if isinstance(value, str):
        return value not in ("0.5", "no normalization")
    return True  # unexpected type -> fail


def main():
    input_file = find_input_file()
    report_file, cleaned_file = derive_output_paths(input_file)
    print(f"Input file: {input_file}")

    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print("ERROR: File has fewer than 2 rows; cannot locate headers.")
        sys.exit(1)

    # ------------------------------------------------------------------ #
    # Step 0 – Header validation
    # ------------------------------------------------------------------ #
    header_row = list(rows[1])  # row 2 (0-indexed row 1)
    actual_headers = [str(h).strip() if h is not None else "" for h in header_row]

    if actual_headers != EXPECTED_HEADERS:
        print("ERROR: Column headers in row 2 do not match the expected headers.")
        print(f"  Expected: {EXPECTED_HEADERS}")
        print(f"  Actual  : {actual_headers}")
        sys.exit(1)

    print("Header validation passed.")

    # ------------------------------------------------------------------ #
    # Prepare data rows
    # ------------------------------------------------------------------ #
    data_rows = rows[2:]  # rows 3+ (0-indexed 2+)
    total = len(data_rows)

    check_names = [
        "Check_Whitespace",
        "Check_SpecialChars",
        "Check_CalPrefix",
        "Check_Concentration",
    ]

    report_headers = EXPECTED_HEADERS + check_names + ["Result"]

    report_rows = []
    cleaned_rows = []

    counts = {name: 0 for name in check_names}
    overall_fail_count = 0

    for row in data_rows:
        row = list(row)
        checks = {}

        # -- Check 1: Whitespace ---------------------------------------- #
        ws_fail = False
        for ci in NAME_COLS:
            if row[ci] is not None and check_whitespace(row[ci]):
                ws_fail = True
        checks["Check_Whitespace"] = "fail" if ws_fail else "pass"

        # Build cleaned row (strip whitespace from name columns)
        cleaned_row = list(row)
        for ci in NAME_COLS:
            if cleaned_row[ci] is not None:
                cleaned_row[ci] = str(cleaned_row[ci]).strip()

        # -- Check 2: Special characters (on trimmed values) ------------- #
        sc_fail = False
        for ci in NAME_COLS:
            val = cleaned_row[ci]
            if val is not None and check_special_chars(val):
                sc_fail = True
        checks["Check_SpecialChars"] = "fail" if sc_fail else "pass"

        # -- Check 3: Cal prefix ---------------------------------------- #
        cal_fail = False
        for ci in NAME_COLS:
            val = cleaned_row[ci]
            if val is not None and check_cal_prefix(val):
                cal_fail = True
        checks["Check_CalPrefix"] = "fail" if cal_fail else "pass"

        # -- Check 4: Concentration ------------------------------------- #
        conc_fail = False
        for ci in CONC_COLS:
            if row[ci] is not None and check_concentration(row[ci]):
                conc_fail = True
        checks["Check_Concentration"] = "fail" if conc_fail else "pass"

        # -- Overall result --------------------------------------------- #
        all_pass = all(v == "pass" for v in checks.values())
        result = "pass" if all_pass else "fail"

        for name in check_names:
            if checks[name] == "fail":
                counts[name] += 1
        if result == "fail":
            overall_fail_count += 1

        report_rows.append(row + [checks[cn] for cn in check_names] + [result])
        cleaned_rows.append(cleaned_row)

    wb.close()

    # ------------------------------------------------------------------ #
    # Write report
    # ------------------------------------------------------------------ #
    rpt_wb = openpyxl.Workbook()
    rpt_ws = rpt_wb.active
    rpt_ws.title = "Validation Report"
    rpt_ws.append(report_headers)
    for r in report_rows:
        rpt_ws.append(r)
    rpt_wb.save(report_file)
    rpt_wb.close()
    print(f"Report written to: {report_file}")

    # ------------------------------------------------------------------ #
    # Write cleaned sample list (full copy with only whitespace trimmed)
    # ------------------------------------------------------------------ #
    shutil.copy2(input_file, cleaned_file)
    cl_wb = openpyxl.load_workbook(cleaned_file)
    cl_ws = cl_wb.worksheets[0]

    for row_idx, cleaned_row in enumerate(cleaned_rows, start=3):  # data starts at row 3
        for ci in NAME_COLS:
            col_letter = openpyxl.utils.get_column_letter(ci + 1)
            cl_ws[f"{col_letter}{row_idx}"] = cleaned_row[ci]

    cl_wb.save(cleaned_file)
    cl_wb.close()
    print(f"Cleaned sample list written to: {cleaned_file}")

    # ------------------------------------------------------------------ #
    # Summary
    # ------------------------------------------------------------------ #
    print(f"\n{'='*50}")
    print(f"Total data rows checked: {total}")
    for name in check_names:
        print(f"  {name}: {counts[name]} row(s) failed")
    print(f"  Overall failures: {overall_fail_count} row(s)")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
